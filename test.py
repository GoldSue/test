import datetime
import openpyxl
from openpyxl import Workbook

class average_time:
    def __init__(self):
        self.ws = None

    def excel(self):
        wb = openpyxl.Workbook()
        self.ws = wb.active
        return self.ws

    def write_excel(self, cell, value):
        self.ws[cell].value = value

    def read_excel(self, cell):
        return self.ws[cell].value

    def public_day(self):
        # 2025年的节假日安排
        public_holidays = [
            "2025-01-01",  # 新年假期
            "2025-01-28",  # 春节假期
            "2025-01-29",
            "2025-01-30",
            "2025-01-31",
            "2025-02-01",
            "2025-02-02",
            "2025-02-03",
            "2025-02-04",
            "2025-04-04",  # 清明节假期
            "2025-04-05",
            "2025-04-06",
            "2025-05-01",  # 劳动节假期
            "2025-05-02",
            "2025-05-03",
            "2025-05-04",
            "2025-05-05",
            "2025-05-31",  # 端午节假期
            "2025-06-01",
            "2025-06-02",
            "2025-10-01",  # 国庆节假期及中秋节假期
            "2025-10-02",
            "2025-10-03",
            "2025-10-04",
            "2025-10-05",
            "2025-10-06",
            "2025-10-07",
            "2025-10-08",
        ]

        # 2025年的调休工作日
        workdays = [
            "2025-01-26",  # 春节前调休
            "2025-02-08",  # 春节后调休
            "2025-04-07",  # 清明节后调休
            "2025-09-28",  # 劳动节后调休
            "2025-10-11",  # 端午节后调休
            "2025-10-09",  # 国庆节假期及中秋节后调休
        ]

        return public_holidays, workdays

    def is_public_holiday(self, date_str, public_holidays):
        return date_str in public_holidays

    def is_workday(self, date_str, workdays):
        return date_str in workdays

    def is_weekend(self, date):
        return date.weekday() >= 5  # 周六是5，周日是6

    def get_all_workdays_of_month(self):
        public_holidays, workdays = self.public_day()
        now = datetime.datetime.now()
        year = now.year
        month = now.month

        workday_list = []

        # 从当月第一天开始
        current_date = datetime.date(year, month, 1)
        # 获取当月最后一天
        if month == 12:
            next_month = datetime.date(year + 1, 1, 1)
        else:
            next_month = datetime.date(year, month + 1, 1)
        end_date = next_month - datetime.timedelta(days=1)

        while current_date <= end_date:
            current_date_str = current_date.strftime("%Y-%m-%d")

            # 检查是否为工作日
            if not self.is_public_holiday(current_date_str, public_holidays) and (not self.is_weekend(current_date) or self.is_workday(current_date_str, workdays)):
                workday_list.append(current_date)

            current_date += datetime.timedelta(days=1)

        return workday_list

    def plan_time(self):
        self.excel()
        col = 2
        for day in self.get_all_workdays_of_month():
            # 写入日期到第一行
            date_cell = self.ws.cell(row=1, column=col)
            date_cell.value = day.strftime("%m-%d")

            # 写入时间安排到第二行和第三行
            cell1 = self.ws.cell(row=2, column=col)
            cell2 = self.ws.cell(row=3, column=col)

            if day.weekday() in (1, 3):  # 周二或周四
                self.write_excel(cell1.coordinate, "8:30")
                self.write_excel(cell2.coordinate, "21:00")
            elif day.weekday() in (0, 2, 4):  # 周一、周三或周五
                self.write_excel(cell1.coordinate, "8:30")
                self.write_excel(cell2.coordinate, "18:00")

            col += 1

        # 保存Excel文件
        self.ws.parent.save("workdays.xlsx")

    def write_workdays_to_excel(self):
        # 获取当月的所有工作日
        workdays_list = self.get_all_workdays_of_month()

        # 创建Excel工作表
        self.excel()

        # 写入工作日到Excel，从B2单元格开始向右依次写入
        col = 2
        for workday in workdays_list:
            cell = self.ws.cell(row=1, column=col)
            cell.value = workday.strftime("%m-%d")
            col += 1

        # 保存Excel文件
        self.ws.parent.save("workdays.xlsx")

# 创建average_time实例
cal = average_time()

# 将当月所有工作日写入Excel文件
cal.plan_time()
print("当月的所有工作日已写入Excel文件：workdays.xlsx")
